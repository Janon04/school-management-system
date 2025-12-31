"""
Views for Student Management
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils.translation import gettext as _
from django.db.models import Q
from apps.accounts.decorators import admin_required, teacher_required
from apps.classes.models import ClassRoom
from .models import Student
from .forms import StudentForm
import os
import pandas as pd
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from io import BytesIO


@login_required
def student_list_view(request):
    """List all active students"""
    query = request.GET.get('q', '')
    class_filter = request.GET.get('class', '')

    students = Student.objects.filter(is_active=True).select_related('user', 'class_assigned', 'parent')
    classes = ClassRoom.objects.filter(is_active=True)

    if query:
        students = students.filter(
            Q(admission_number__icontains=query) |
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query)
        )

    if class_filter:
        students = students.filter(class_assigned_id=class_filter)

    male_count = students.filter(gender='M').count()
    female_count = students.filter(gender='F').count()

    context = {
        'students': students,
        'classes': classes,
        'query': query,
        'class_filter': class_filter,
        'male_count': male_count,
        'female_count': female_count,
    }
    return render(request, 'students/student_list.html', context)


@login_required
def student_detail_view(request, pk):
    """View student details"""
    student = get_object_or_404(Student, pk=pk)
    
    # Check permission - students can only view their own profile
    if request.user.is_student and student.user != request.user:
        messages.error(request, _('You can only view your own profile.'))
        return redirect('dashboard')
    
    context = {
        'student': student,
    }
    return render(request, 'students/student_detail.html', context)


@login_required
@admin_required
def student_create_view(request):
    """Create new student (admin only)"""
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            student = form.save()
            messages.success(
                request,
                _('Student %(name)s (%(admission)s) created successfully!') % {
                    'name': student.user.get_full_name(),
                    'admission': student.admission_number
                }
            )
            return redirect('students:student_detail', pk=student.pk)
        else:
            messages.error(request, _('Please correct the errors below.'))
    else:
        form = StudentForm()
    
    context = {
        'form': form,
        'title': _('Add New Student'),
        'button_text': _('Create Student'),
    }
    return render(request, 'students/student_form.html', context)


@login_required
@admin_required
def student_update_view(request, pk):
    """Update student information (admin only)"""
    student = get_object_or_404(Student, pk=pk)
    
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            student = form.save()
            messages.success(
                request,
                _('Student %(name)s updated successfully!') % {
                    'name': student.user.get_full_name()
                }
            )
            return redirect('students:student_detail', pk=student.pk)
        else:
            messages.error(request, _('Please correct the errors below.'))
    else:
        form = StudentForm(instance=student)
    
    context = {
        'form': form,
        'student': student,
        'title': _('Edit Student - %(name)s') % {'name': student.user.get_full_name()},
        'button_text': _('Update Student'),
    }
    return render(request, 'students/student_form.html', context)


@login_required
@admin_required
def student_delete_view(request, pk):
    """Deactivate student (admin only)"""
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        student.is_active = False
        student.save()
        messages.success(request, _('Student %(admission)s deactivated.') % {'admission': student.admission_number})
        return redirect('students:student_list')
    
    return render(request, 'students/student_confirm_delete.html', {'student': student})


@login_required
@admin_required
def student_import_view(request):
    """Import students from Excel template (admin only)"""
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'students'))
        filename = fs.save(file.name, file)
        file_path = fs.path(filename)
        try:
            df = pd.read_excel(file_path)
            required_columns = [
                'Registration Number', 'First Name', 'Middle Name', 'Last Name', 'Gender',
                'Date of Birth', 'Class'
            ]
            for col in required_columns:
                if col not in df.columns:
                    messages.error(request, f"Missing required column: {col}")
                    return redirect('students:student_import')
            from apps.classes.models import ClassRoom
            from apps.parents.models import Parent
            from django.contrib.auth import get_user_model
            User = get_user_model()
            imported, errors = 0, 0
            for _, row in df.iterrows():
                try:
                    user, created = User.objects.get_or_create(
                        username=row['Registration Number'],
                        defaults={
                            'first_name': row['First Name'],
                            'last_name': row['Last Name'],
                            'role': 'STUDENT',
                        }
                    )
                    class_obj = ClassRoom.objects.filter(name=row['Class']).first()
                    if not class_obj:
                        errors += 1
                        continue
                    parent = None
                    if pd.notna(row.get('Guardian Full Name')):
                        parent, _ = Parent.objects.get_or_create(
                            user__first_name=row['Guardian Full Name'].split()[0],
                            user__last_name=' '.join(row['Guardian Full Name'].split()[1:]) if len(row['Guardian Full Name'].split()) > 1 else '',
                        )
                    Student.objects.get_or_create(
                        user=user,
                        defaults={
                            'admission_number': row['Registration Number'],
                            'middle_name': row.get('Middle Name', ''),
                            'gender': row['Gender'],
                            'date_of_birth': row['Date of Birth'],
                            'class_assigned': class_obj,
                            'parent': parent,
                            'emergency_contact_name': row.get('emergency_contact', ''),
                            'previous_school': row.get('previous_school', ''),
                            'religion': row.get('religion', ''),
                            'roll_number': row.get('roll_number', ''),
                            'medical_conditions': row.get('medical_conditions', ''),
                            'nationality': row.get('Nationality', ''),
                            'academic_year': class_obj.academic_year if class_obj else None,
                            'admission_date': row.get('Enrollment Date', None),
                        }
                    )
                    imported += 1
                except Exception as e:
                    errors += 1
            messages.success(request, f"Imported {imported} students. {errors} errors.")
            return redirect('students:student_list')
        except Exception as e:
            messages.error(request, f"Import failed: {str(e)}")
            return redirect('students:student_import')
    return render(request, 'students/student_import.html')


def student_template_download_view(request):
    """Dynamically generate and return a student import Excel template with validations and two sample students."""
    from apps.classes.models import ClassRoom, AcademicYear
    from django.utils import timezone
    from .models import Student

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Student Import Template'

    columns = [
        ("Registration Number", True, None),
        ("First Name", True, None),
        ("Middle Name", False, None),
        ("Last Name", True, None),
        ("Gender", True, ['M', 'F']),
        ("Date of Birth", True, None),
        ("Nationality", True, None),
        ("Class", True, list(ClassRoom.objects.values_list('name', flat=True))),
        ("Stream/Combination", False, None),
        ("Academic Year", True, list(AcademicYear.objects.values_list('name', flat=True))),
        ("Enrollment Date", True, None),
        ("Student Status", True, ['Active', 'Inactive', 'Graduated', 'Transferred']),
        ("Village", False, None),
        ("Cell", False, None),
        ("Sector", False, None),
        ("District", False, None),
        ("Guardian Full Name", False, None),
        ("Relationship", False, None),
        ("Guardian Phone", False, None),
        ("Guardian Email", False, None),
        ("Medical Conditions", False, None),
        ("Special Needs", False, None),
        ("Remarks/Notes", False, None),
    ]
    ws.append([col[0] for col in columns])
    # Add two sample students from DB
    sample_students = Student.objects.filter(is_active=True)[:2]
    for s in sample_students:
        ws.append([
            s.admission_number,
            s.user.first_name,
            s.middle_name,
            s.user.last_name,
            s.gender,
            s.date_of_birth.isoformat() if s.date_of_birth else '',
            s.nationality,
            s.class_assigned.name if s.class_assigned else '',
            getattr(s.class_assigned, 'stream', '') if s.class_assigned else '',
            s.academic_year.name if s.academic_year else '',
            s.admission_date.isoformat() if s.admission_date else '',
            'Active' if s.is_active else 'Inactive',
            '', '', '', '',
            s.parent.user.get_full_name() if s.parent and s.parent.user else '',
            '', '', '',
            s.medical_conditions,
            '', ''
        ])
    # Add data validation for columns with choices
    for idx, (header, required, choices) in enumerate(columns, 1):
        col_letter = get_column_letter(idx)
        if choices:
            dv = DataValidation(type="list", formula1='"' + ','.join(choices) + '"', allow_blank=not required)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}1048576")
        ws[f"{col_letter}1"].font = openpyxl.styles.Font(bold=True, color="FF0000" if required else "000000")
    for idx, (header, *_rest) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = max(15, len(header) + 2)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=student_import_template.xlsx'
    return response
